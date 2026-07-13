"""Peru vehicle plate verification Streamlit app."""

import os
import io
import time
from datetime import datetime

import streamlit as st
import requests
import openpyxl
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def load_dotenv(path='.env'):
    """Minimal .env loader (no external dependency)."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, _, value = line.partition('=')
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_dotenv()

# Configuration
TOKEN = os.environ.get('VERIFIK_TOKEN', '')
AUTH_SCHEME = os.environ.get('VERIFIK_AUTH_SCHEME', 'Bearer')
REQUEST_TIMEOUT = 30
DELAY_BETWEEN_CALLS = 0.5  # seconds

COUNTRIES = {
    'Argentina': 'ar',
    'Bolivia': 'bo',
    'Brazil': 'br',
    'Chile': 'cl',
    'Colombia': 'co',
    'Costa Rica': 'cr',
    'Ecuador': 'ec',
    'Mexico': 'mx',
    'Paraguay': 'py',
    'Peru': 'pe',
    'United States': 'us',
}

def get_api_url(country_code):
    return f'https://api.verifik.co/v2/{country_code}/vehiculo/placa'

FIELDS = [
    ('Use', ['use', 'tipoUso']),
    ('Type', ['type', 'clase', 'categoria']),
    ('Brand', ['brand', 'marca']),
    ('Model', ['model', 'modelo']),
    ('Year', ['year', 'anoFabricacion', 'ano']),
    ('Engine Serial', ['engineSerial', 'numeroMotor']),
    ('Chassis Serial', ['chasisSerial', 'numeroChasis']),
    ('Seats', ['seats', 'asientos']),
    ('Serial', ['serial', 'serie']),
    ('Valid Format', ['validFormat']),
]
HEADERS = ['Result'] + [h for h, _ in FIELDS] + ['Verified Date']


@st.cache_resource
def get_session():
    """Create a requests session with retry strategy."""
    session = requests.Session()

    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=['GET']
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount('https://', adapter)
    session.mount('http://', adapter)

    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json',
        'Authorization': f'{AUTH_SCHEME} {TOKEN}',
    })

    return session


def verify_plate(session, plate_number, api_url, debug=False):
    """Query the Verifik API for a single plate."""
    row = {h: '' for h in HEADERS}

    if debug:
        st.write(f"**Debug:** Sending plate: `{plate_number}`")

    try:
        resp = session.get(
            api_url,
            params={'plate': plate_number},
            timeout=REQUEST_TIMEOUT,
            verify=False,
        )

        if debug:
            st.write(f"**Debug:** API Status: {resp.status_code}")
    except requests.exceptions.Timeout:
        row['Result'] = 'Timeout'
        return row
    except requests.exceptions.ConnectionError as e:
        error_msg = str(e)
        if 'Max retries exceeded' in error_msg:
            row['Result'] = 'Connection failed'
        else:
            row['Result'] = 'Connection error'
        return row
    except requests.RequestException as e:
        row['Result'] = f'Error: {str(e)[:40]}'
        return row

    if resp.status_code == 200:
        res_json = resp.json() or {}
        data = res_json.get('data', {})

        if debug:
            st.write(f"**Debug:** API Response: {data}")

        for header, keys in FIELDS:
            value = None
            for key in keys:
                if key in data:
                    value = data[key]
                    break
            row[header] = '' if value is None else str(value)

        brand = row.get('Brand', '')
        model = row.get('Model', '')
        row['Result'] = 'Found' if (brand or model) else 'No data'

    elif resp.status_code == 404:
        row['Result'] = 'Not found'
    else:
        try:
            detail = resp.json().get('message') or resp.json().get('code') or resp.text[:40]
        except ValueError:
            detail = resp.text[:40]
        row['Result'] = f'HTTP {resp.status_code}'

    return row


def create_excel_output(rows, plates, uploaded_file=None):
    """Create or append results to Excel file."""
    if uploaded_file:
        # Load existing workbook and add results to new sheet
        wb = openpyxl.load_workbook(uploaded_file)
    else:
        # Create new workbook
        wb = openpyxl.Workbook()

    ws = wb.create_sheet("Verification Results")

    # Write plate numbers in column A
    ws.cell(row=1, column=1, value="Plate Number")
    for idx, plate in enumerate(plates, start=2):
        ws.cell(row=idx, column=1, value=plate)

    # Write headers starting from column B
    for col, header in enumerate(HEADERS, start=2):
        ws.cell(row=1, column=col, value=header)

    # Write data
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for idx, row in enumerate(rows, start=2):
        row['Verified Date'] = now
        for col, header in enumerate(HEADERS, start=2):
            ws.cell(row=idx, column=col, value=row.get(header, ''))

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


st.set_page_config(page_title="Vehicle Plate Verification", layout="wide", initial_sidebar_state="collapsed")

st.title("🚗 Vehicle Plate Verification")
st.caption("Verify vehicle details using the Verifik API")

if not TOKEN:
    st.error("❌ VERIFIK_TOKEN environment variable not set. Please configure it first.")
    st.stop()

# Country selection
selected_country = st.selectbox(
    "Select Country",
    options=list(COUNTRIES.keys()),
    index=list(COUNTRIES.keys()).index('Peru'),
    help="Choose the country to verify plates for"
)
country_code = COUNTRIES[selected_country]
api_url = get_api_url(country_code)

# Input method selection
input_method = st.radio("How do you want to input plates?", ["Single plate", "Upload Excel file"], horizontal=True)

plates_to_verify = []
source_file = None

if input_method == "Single plate":
    plate_input = st.text_input("Enter plate number (e.g., F2X112)", placeholder="ABC123").strip().upper()
    if plate_input:
        plates_to_verify = [plate_input.replace("-", "")]
else:
    uploaded_file = st.file_uploader("Upload Excel file with plates", type=['xlsx'])
    if uploaded_file:
        try:
            source_file = uploaded_file
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            # Show preview and let user select column
            preview_data = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
            st.write("**Preview (first 5 rows):**")
            st.dataframe(preview_data)

            col_letter = st.selectbox("Select column with plate numbers",
                                     ["A", "B", "C", "D", "E", "F"],
                                     help="Choose the column containing plate numbers")
            col_index = ord(col_letter) - ord('A')  # Convert A->0, B->1, C->2, etc.

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row and len(row) > col_index and row[col_index]:
                    plate = str(row[col_index]).strip().replace("-", "").upper()
                    if plate:
                        plates_to_verify.append(plate)
            wb.close()
            st.success(f"✅ Loaded {len(plates_to_verify)} plates from column {col_letter}")
        except Exception as e:
            st.error(f"Error reading file: {e}")
            st.stop()

if plates_to_verify:
    if st.button("🔍 Verify Plates", type="primary"):
        session = get_session()
        results = []

        progress_bar = st.progress(0)
        status_text = st.empty()

        for i, plate in enumerate(plates_to_verify):
            status_text.text(f"Verifying {i+1}/{len(plates_to_verify)}: {plate}")
            is_debug = len(plates_to_verify) == 1
            result = verify_plate(session, plate, api_url, debug=is_debug)
            results.append(result)
            progress_bar.progress((i + 1) / len(plates_to_verify))

            if i < len(plates_to_verify) - 1:
                time.sleep(DELAY_BETWEEN_CALLS)

        status_text.empty()
        progress_bar.empty()

        # Display results
        st.subheader("Results")

        # Create a display dataframe with selected columns
        display_df = []
        for plate, row in zip(plates_to_verify, results):
            display_row = {
                "Plate": plate,
                "Status": row['Result'],
                "Brand": row.get('Brand', ''),
                "Model": row.get('Model', ''),
                "Year": row.get('Year', ''),
                "Type": row.get('Type', ''),
                "Seats": row.get('Seats', ''),
                "Engine Serial": row.get('Engine Serial', ''),
                "Chassis Serial": row.get('Chassis Serial', ''),
            }
            display_df.append(display_row)

        st.dataframe(display_df, use_container_width=True, height=25 + (len(display_df) * 35))

        # Summary stats
        col1, col2, col3 = st.columns(3)
        found = sum(1 for r in results if r['Result'] == 'Found')
        not_found = sum(1 for r in results if r['Result'] == 'Not found')
        errors = sum(1 for r in results if 'Error' in r['Result'] or 'failed' in r['Result'].lower())

        col1.metric("Found", found)
        col2.metric("Not Found", not_found)
        col3.metric("Errors", errors)

        # Download button
        excel_file = create_excel_output(results, plates_to_verify, uploaded_file=source_file)

        if source_file:
            file_name = f"verified_{source_file.name.replace('.xlsx', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            file_name = f"verification_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.download_button(
            label="📥 Download Excel Results",
            data=excel_file,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
